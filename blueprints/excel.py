import os.path

from flask import Blueprint, jsonify
from openpyxl import load_workbook
from pathlib import Path

from app import database

from models.user import User
from validators.user_validation import UserValidation

excel = Blueprint("excel_blueprint", __name__)


@excel.route("/read", methods=["POST"])
def read_xl_file():
    """
    function to read the Excel file rows and columns, validate them and insert them into the database.
    :return: (json) message stating how many rows in the Excel have been parsed, or an error message.
    """

    file_path = os.path.join("static/Library_register_data.xlsx")

    if not os.path.exists(file_path):
        return jsonify({"error": f"The file of {file_path} doesn't exist."}), 404

    library_register = load_workbook(file_path)
    sheet = library_register.active
    row_counter = 0
    # improve this file handling by using pandas to read the Excel in chunks, to prevent loading all into memory

    # loop over each row in the Excel file
    for row in range(2, sheet.max_row + 1):
        row_info = []
        # loop over each column in the Excel file
        for column in range(1, sheet.max_column + 1):

            # check if the cell value is not null and append it to the row
            cell = sheet.cell(row=row, column=column)
            if cell.value is not None:
                row_info.append(cell.value)

        # check if the row has data and try to add the 'validated' person to the database
        if row_info:
            try:
                add_person(validate_data(row_info))
                row_counter += 1
            except ValueError as error:
                return jsonify({"error": f"{error}"}), 400

    return jsonify({"message": f"{row_counter} users parsed and added to the DB."}), 200


def add_person(person: UserValidation):
    """Add a 'validated' person to the database.
    :param person A object of userValidation that can be added to the database
    """
    new_user = User(
        first_name=person.name["first_name"],
        last_name=person.name["last_name"],
        date_of_birth=person.date_of_birth,
        time_in=person.time_in,
        membership_no=person.membership_no,
        valid_from=person.valid_from,
        valid_to=person.valid_to,
        gender=person.gender,
        researcher=person.researcher
    )

    database.session.add(new_user)
    database.session.commit()


def validate_data(row) -> UserValidation:
    """
    pass the incoming row from the Excel off to pydantic user validation/serialisation.
    :param row: the incoming row to be validated
    :return: the validated / serialised user.
    """
    user = UserValidation(
        name=row[0],
        date_of_birth=row[1],
        time_in=row[2],
        membership_no=row[3],
        valid_from=row[4],
        valid_to=row[5],
        gender=row[6],
        researcher=row[7],
    )

    user.model_validate(user)
    return user
